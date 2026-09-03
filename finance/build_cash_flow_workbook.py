import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment

OUT = "/tmp/claude-0/-home-user-content-calendar-demo/5754c680-8ffc-51b0-a0da-15c1d7790787/scratchpad/Cash_Flow_Statement_August_2026.xlsx"

# ---------- styling ----------
FONT = "Arial"
NAVY = "1F3864"; MID = "2F5496"; LIGHT = "D9E2F3"; PALE = "F2F5FB"; GREY = "F2F2F2"
GREEN = "E2EFDA"; RED = "FCE4E4"; AMBER = "FFF2CC"; YELLOW = "FFFF00"
INPUT_BLUE = "0000FF"; LINK_GREEN = "008000"

thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color=NAVY)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP2 = Border(top=Side(style="double", color=NAVY), bottom=Side(style="double", color=NAVY))

TRY = '"₺"#,##0.00;[Red]("₺"#,##0.00);"-"'
NUM2 = '#,##0.00;[Red](#,##0.00);"-"'
FX = '0.0000'
PCT = '0.0%;[Red]-0.0%;"-"'
DATE = 'dd-mmm-yyyy'
INT = '0;[Red]-0;"-"'

def f(bold=False, color="000000", size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)

def fill(c): return PatternFill("solid", start_color=c, end_color=c)

def hdr(ws, row, col, text, width=None, wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = fill(NAVY); c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if width: ws.column_dimensions[L(col)].width = width
    return c

def title(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    a = ws["A1"]; a.value = text; a.font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
    a.fill = fill(NAVY); a.alignment = Alignment(vertical="center", indent=1)
    b = ws["A2"]; b.value = sub; b.font = Font(name=FONT, italic=True, size=10, color="FFFFFF")
    b.fill = fill(MID); b.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30; ws.row_dimensions[2].height = 18
    ws.sheet_view.showGridLines = False

def setcell(ws, ref, value, fmt=None, font=None, fl=None, align=None, border=True):
    c = ws[ref]; c.value = value
    if fmt: c.number_format = fmt
    c.font = font or f()
    if fl: c.fill = fill(fl)
    if align: c.alignment = align
    if border: c.border = BORDER
    return c

N_ROWS = 40   # data rows per register
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFTC = Alignment(horizontal="left", vertical="center", indent=1)

wb = Workbook()
dash = wb.active; dash.title = "Dashboard"
ar = wb.create_sheet("Receivables")
ap = wb.create_sheet("Payables")
hq = wb.create_sheet("UK HQ Invoices")
bk = wb.create_sheet("Bank Balances")

# Key dashboard input references (absolute)
REPORT_DATE = "Dashboard!$C$8"
W_START = ["Dashboard!$C$10", "Dashboard!$C$11", "Dashboard!$C$12", "Dashboard!$C$13", "Dashboard!$C$14"]
FX_TBL = "Dashboard!$F$9:$G$12"      # Currency | Rate to TRY

# =====================================================================
# RECEIVABLES
# =====================================================================
title(ar, "ACCOUNTS RECEIVABLE – CUSTOMER COLLECTIONS", "Amounts in Turkish Lira (₺). Blue cells are inputs; black cells are formulas – do not overwrite.", 17)
ar_cols = ["No", "Customer", "Invoice No", "Invoice Date", "Due Date", "Invoice Amount (₺)",
           "Week 1", "Week 2", "Week 3", "Week 4", "Week 5",
           "Total Collected (₺)", "Remaining (₺)", "Collected %", "Days Overdue", "Status", "Notes"]
ar_w = [5, 28, 14, 13, 13, 18, 15, 15, 15, 15, 15, 18, 18, 11, 11, 12, 30]
HR = 5  # header row
ar.merge_cells(start_row=4, start_column=7, end_row=4, end_column=11)
c = ar.cell(row=4, column=7, value="COLLECTIONS RECEIVED BY WEEK (₺)  –  week start dates below")
c.font = f(True, "FFFFFF"); c.fill = fill(MID); c.alignment = CENTER
for i, (h, w) in enumerate(zip(ar_cols, ar_w), 1):
    hdr(ar, HR, i, h, w)
# week date row under headers
ar.row_dimensions[HR].height = 32
for i, ws_ref in enumerate(W_START):
    cc = ar.cell(row=HR+1, column=7+i, value=f"={ws_ref}")
    cc.number_format = DATE; cc.font = f(True, LINK_GREEN, 9); cc.fill = fill(LIGHT); cc.alignment = CENTER; cc.border = BORDER
for col in range(1, 18):
    if col not in range(7, 12):
        cc = ar.cell(row=HR+1, column=col); cc.fill = fill(LIGHT); cc.border = BORDER
ar.cell(row=HR+1, column=2, value="Week starting →").font = f(True, NAVY, 9)
ar.cell(row=HR+1, column=2).alignment = RIGHT

AR_FIRST, AR_LAST = HR+2, HR+1+N_ROWS
ar_sample = [
    ("Anadolu Retail A.Ş.", "INV-2026-081", dt.date(2026,7,15), dt.date(2026,8,14), 185000, [60000, 60000, 0, 0, 0]),
    ("Marmara Lojistik Ltd.", "INV-2026-084", dt.date(2026,7,28), dt.date(2026,8,27), 92500, [0, 0, 92500, 0, 0]),
    ("Ege Gıda San. Tic.", "INV-2026-087", dt.date(2026,8,3), dt.date(2026,9,2), 240000, [0, 0, 0, 0, 0]),
]
for r in range(AR_FIRST, AR_LAST+1):
    i = r - AR_FIRST
    zebra = PALE if i % 2 else None
    setcell(ar, f"A{r}", f'=IF(B{r}="","",ROW()-{AR_FIRST-1})', INT, f(color="000000"), zebra, CENTER)
    for col, fmt in [("B", None), ("C", None), ("D", DATE), ("E", DATE), ("F", TRY), ("Q", None)]:
        setcell(ar, f"{col}{r}", None, fmt, f(color=INPUT_BLUE), zebra, CENTER if col in "DE" else (RIGHT if col == "F" else LEFTC))
    for col in "GHIJK":
        setcell(ar, f"{col}{r}", None, TRY, f(color=INPUT_BLUE), zebra, RIGHT)
    setcell(ar, f"L{r}", f'=IF(B{r}="","",SUM(G{r}:K{r}))', TRY, f(), zebra, RIGHT)
    setcell(ar, f"M{r}", f'=IF(B{r}="","",F{r}-L{r})', TRY, f(True), zebra, RIGHT)
    setcell(ar, f"N{r}", f'=IF(OR(B{r}="",F{r}=0),"",L{r}/F{r})', PCT, f(), zebra, RIGHT)
    setcell(ar, f"O{r}", f'=IF(OR(B{r}="",E{r}="",M{r}<=0),"",MAX(0,{REPORT_DATE}-E{r}))', INT, f(), zebra, CENTER)
    setcell(ar, f"P{r}", f'=IF(B{r}="","",IF(M{r}<=0,"Collected",IF(AND(O{r}<>"",O{r}>0),"Overdue","Open")))', None, f(True), zebra, CENTER)
    if i < len(ar_sample):
        s = ar_sample[i]
        ar[f"B{r}"] = s[0]; ar[f"C{r}"] = s[1]; ar[f"D{r}"] = s[2]; ar[f"E{r}"] = s[3]; ar[f"F{r}"] = s[4]
        for k, v in enumerate(s[5]):
            ar.cell(row=r, column=7+k, value=v if v else None)
        ar[f"Q{r}"] = "Sample row – replace with live data"
TR = AR_LAST + 1
setcell(ar, f"B{TR}", "TOTAL", None, f(True, NAVY, 11), LIGHT, LEFTC)
for col in "ACDEQ":
    setcell(ar, f"{col}{TR}", None, None, f(), LIGHT)
for col in "FGHIJKLM":
    setcell(ar, f"{col}{TR}", f"=SUM({col}{AR_FIRST}:{col}{AR_LAST})", TRY, f(True, NAVY, 11), LIGHT, RIGHT)
setcell(ar, f"N{TR}", f'=IF(F{TR}=0,"",L{TR}/F{TR})', PCT, f(True, NAVY, 11), LIGHT, RIGHT)
setcell(ar, f"O{TR}", f'=COUNTIF(P{AR_FIRST}:P{AR_LAST},"Overdue")', INT, f(True, "C00000", 11), LIGHT, CENTER)
ar[f"O{TR}"].comment = Comment("Number of overdue invoices", "Finance")
setcell(ar, f"P{TR}", "overdue", None, f(True, "C00000", 9), LIGHT, CENTER)
for col in range(1, 18):
    ar.cell(row=TR, column=col).border = TOP2
# checks
setcell(ar, f"B{TR+2}", "Control: Invoice total – Collected – Remaining (must be 0)", None, f(italic=True, size=9), border=False)
setcell(ar, f"F{TR+2}", f"=F{TR}-L{TR}-M{TR}", TRY, f(True, size=9), None, RIGHT)
ar.freeze_panes = f"C{AR_FIRST}"
ar.auto_filter.ref = f"A{HR}:Q{AR_LAST}"
dv = DataValidation(type="list", formula1='"Open,Overdue,Collected"', allow_blank=True); ar.add_data_validation(dv)
dvd = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True,
                     error="Enter a valid date", errorTitle="Date"); ar.add_data_validation(dvd)
dvd.add(f"D{AR_FIRST}:E{AR_LAST}")
rng = f"P{AR_FIRST}:P{AR_LAST}"
ar.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Overdue"'], fill=fill(RED), font=Font(name=FONT, bold=True, color="9C0006")))
ar.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Collected"'], fill=fill(GREEN), font=Font(name=FONT, bold=True, color="006100")))
ar.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Open"'], fill=fill(AMBER), font=Font(name=FONT, bold=True, color="9C5700")))
ar.conditional_formatting.add(f"M{AR_FIRST}:M{AR_LAST}", CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED)))
ar.conditional_formatting.add(f"N{AR_FIRST}:N{AR_LAST}", FormulaRule(formula=[f'AND(N{AR_FIRST}<>"",N{AR_FIRST}>1)'], fill=fill(RED)))

# =====================================================================
# PAYABLES
# =====================================================================
title(ap, "ACCOUNTS PAYABLE – SCHEDULED PAYMENTS", "Amounts in Turkish Lira (₺). Blue cells are inputs; black cells are formulas – do not overwrite.", 18)
ap_cols = ["No", "Supplier / Payee", "Category", "Invoice No", "Invoice Date", "Due Date", "Invoice Amount (₺)",
           "Week 1", "Week 2", "Week 3", "Week 4", "Week 5",
           "Total Paid (₺)", "Remaining (₺)", "Paid %", "Days Overdue", "Status", "Notes"]
ap_w = [5, 28, 16, 14, 13, 13, 18, 15, 15, 15, 15, 15, 18, 18, 11, 11, 12, 30]
ap.merge_cells(start_row=4, start_column=8, end_row=4, end_column=12)
c = ap.cell(row=4, column=8, value="PAYMENTS PLANNED / MADE BY WEEK (₺)  –  week start dates below")
c.font = f(True, "FFFFFF"); c.fill = fill(MID); c.alignment = CENTER
for i, (h, w) in enumerate(zip(ap_cols, ap_w), 1):
    hdr(ap, HR, i, h, w)
ap.row_dimensions[HR].height = 32
for i, ws_ref in enumerate(W_START):
    cc = ap.cell(row=HR+1, column=8+i, value=f"={ws_ref}")
    cc.number_format = DATE; cc.font = f(True, LINK_GREEN, 9); cc.fill = fill(LIGHT); cc.alignment = CENTER; cc.border = BORDER
for col in range(1, 19):
    if col not in range(8, 13):
        cc = ap.cell(row=HR+1, column=col); cc.fill = fill(LIGHT); cc.border = BORDER
ap.cell(row=HR+1, column=2, value="Week starting →").font = f(True, NAVY, 9)
ap.cell(row=HR+1, column=2).alignment = RIGHT
AP_FIRST, AP_LAST = HR+2, HR+1+N_ROWS
CATS = ["Payroll", "Rent & Utilities", "Suppliers / COGS", "Taxes & Social Security", "Marketing", "IT & Software", "Professional Fees", "Loan / Finance", "Other"]
ap_sample = [
    ("Staff Payroll – August", "Payroll", "PAY-08-2026", dt.date(2026,8,25), dt.date(2026,8,31), 320000, [0, 0, 0, 320000, 0]),
    ("İstanbul Office Rent", "Rent & Utilities", "RNT-2026-08", dt.date(2026,8,1), dt.date(2026,8,5), 85000, [85000, 0, 0, 0, 0]),
    ("Karadeniz Ambalaj Ltd.", "Suppliers / COGS", "SUP-4471", dt.date(2026,7,20), dt.date(2026,8,19), 64300, [0, 0, 30000, 0, 0]),
]
for r in range(AP_FIRST, AP_LAST+1):
    i = r - AP_FIRST
    zebra = PALE if i % 2 else None
    setcell(ap, f"A{r}", f'=IF(B{r}="","",ROW()-{AP_FIRST-1})', INT, f(), zebra, CENTER)
    for col, fmt in [("B", None), ("C", None), ("D", None), ("E", DATE), ("F", DATE), ("G", TRY), ("R", None)]:
        setcell(ap, f"{col}{r}", None, fmt, f(color=INPUT_BLUE), zebra, CENTER if col in "EF" else (RIGHT if col == "G" else LEFTC))
    for col in "HIJKL":
        setcell(ap, f"{col}{r}", None, TRY, f(color=INPUT_BLUE), zebra, RIGHT)
    setcell(ap, f"M{r}", f'=IF(B{r}="","",SUM(H{r}:L{r}))', TRY, f(), zebra, RIGHT)
    setcell(ap, f"N{r}", f'=IF(B{r}="","",G{r}-M{r})', TRY, f(True), zebra, RIGHT)
    setcell(ap, f"O{r}", f'=IF(OR(B{r}="",G{r}=0),"",M{r}/G{r})', PCT, f(), zebra, RIGHT)
    setcell(ap, f"P{r}", f'=IF(OR(B{r}="",F{r}="",N{r}<=0),"",MAX(0,{REPORT_DATE}-F{r}))', INT, f(), zebra, CENTER)
    setcell(ap, f"Q{r}", f'=IF(B{r}="","",IF(N{r}<=0,"Paid",IF(AND(P{r}<>"",P{r}>0),"Overdue","Scheduled")))', None, f(True), zebra, CENTER)
    if i < len(ap_sample):
        s = ap_sample[i]
        ap[f"B{r}"] = s[0]; ap[f"C{r}"] = s[1]; ap[f"D{r}"] = s[2]; ap[f"E{r}"] = s[3]; ap[f"F{r}"] = s[4]; ap[f"G{r}"] = s[5]
        for k, v in enumerate(s[6]):
            ap.cell(row=r, column=8+k, value=v if v else None)
        ap[f"R{r}"] = "Sample row – replace with live data"
TP = AP_LAST + 1
setcell(ap, f"B{TP}", "TOTAL", None, f(True, NAVY, 11), LIGHT, LEFTC)
for col in "ACDEFR":
    setcell(ap, f"{col}{TP}", None, None, f(), LIGHT)
for col in "GHIJKLMN":
    setcell(ap, f"{col}{TP}", f"=SUM({col}{AP_FIRST}:{col}{AP_LAST})", TRY, f(True, NAVY, 11), LIGHT, RIGHT)
setcell(ap, f"O{TP}", f'=IF(G{TP}=0,"",M{TP}/G{TP})', PCT, f(True, NAVY, 11), LIGHT, RIGHT)
setcell(ap, f"P{TP}", f'=COUNTIF(Q{AP_FIRST}:Q{AP_LAST},"Overdue")', INT, f(True, "C00000", 11), LIGHT, CENTER)
setcell(ap, f"Q{TP}", "overdue", None, f(True, "C00000", 9), LIGHT, CENTER)
for col in range(1, 19):
    ap.cell(row=TP, column=col).border = TOP2
setcell(ap, f"B{TP+2}", "Control: Invoice total – Paid – Remaining (must be 0)", None, f(italic=True, size=9), border=False)
setcell(ap, f"G{TP+2}", f"=G{TP}-M{TP}-N{TP}", TRY, f(True, size=9), None, RIGHT)
ap.freeze_panes = f"C{AP_FIRST}"
ap.auto_filter.ref = f"A{HR}:R{AP_LAST}"
dvc = DataValidation(type="list", formula1='"' + ",".join(CATS) + '"', allow_blank=True, error="Choose a category from the list", errorTitle="Category")
ap.add_data_validation(dvc); dvc.add(f"C{AP_FIRST}:C{AP_LAST}")
dvd2 = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True); ap.add_data_validation(dvd2); dvd2.add(f"E{AP_FIRST}:F{AP_LAST}")
rng = f"Q{AP_FIRST}:Q{AP_LAST}"
ap.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Overdue"'], fill=fill(RED), font=Font(name=FONT, bold=True, color="9C0006")))
ap.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Paid"'], fill=fill(GREEN), font=Font(name=FONT, bold=True, color="006100")))
ap.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Scheduled"'], fill=fill(AMBER), font=Font(name=FONT, bold=True, color="9C5700")))
ap.conditional_formatting.add(f"N{AP_FIRST}:N{AP_LAST}", CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED)))

# =====================================================================
# UK HQ INVOICES
# =====================================================================
title(hq, "UK HQ – OUTSTANDING INTERCOMPANY INVOICES", "Invoices received from UK Head Office not yet settled. Foreign currency amounts convert to ₺ using the FX rates on the Dashboard.", 20)
hq_cols = ["No", "HQ Invoice No", "Invoice Date", "Due Date", "Description / Service", "Currency", "Invoice Amount (FCY)", "FX Rate (to ₺)", "Invoice Amount (₺)",
           "Week 1", "Week 2", "Week 3", "Week 4", "Week 5",
           "Total Paid (₺)", "Remaining (₺)", "Days Overdue", "Aging Bucket", "Status", "Notes"]
hq_w = [5, 15, 13, 13, 32, 10, 18, 12, 18, 15, 15, 15, 15, 15, 18, 18, 11, 13, 12, 30]
hq.merge_cells(start_row=4, start_column=10, end_row=4, end_column=14)
c = hq.cell(row=4, column=10, value="SETTLEMENTS PLANNED / MADE BY WEEK (₺)  –  week start dates below")
c.font = f(True, "FFFFFF"); c.fill = fill(MID); c.alignment = CENTER
for i, (h, w) in enumerate(zip(hq_cols, hq_w), 1):
    hdr(hq, HR, i, h, w)
hq.row_dimensions[HR].height = 32
for i, ws_ref in enumerate(W_START):
    cc = hq.cell(row=HR+1, column=10+i, value=f"={ws_ref}")
    cc.number_format = DATE; cc.font = f(True, LINK_GREEN, 9); cc.fill = fill(LIGHT); cc.alignment = CENTER; cc.border = BORDER
for col in range(1, 21):
    if col not in range(10, 15):
        cc = hq.cell(row=HR+1, column=col); cc.fill = fill(LIGHT); cc.border = BORDER
hq.cell(row=HR+1, column=5, value="Week starting →").font = f(True, NAVY, 9)
hq.cell(row=HR+1, column=5).alignment = RIGHT
HQ_FIRST, HQ_LAST = HR+2, HR+1+N_ROWS
hq_sample = [
    ("UKHQ-2026-0412", dt.date(2026,5,30), dt.date(2026,6,29), "Management fee – Q2 2026", "GBP", 12500, [0,0,0,0,0]),
    ("UKHQ-2026-0498", dt.date(2026,6,30), dt.date(2026,7,30), "Software licences recharge", "GBP", 4200, [0,0,0,0,0]),
    ("UKHQ-2026-0533", dt.date(2026,7,31), dt.date(2026,8,30), "Group insurance recharge", "EUR", 3800, [0,0,0,0,0]),
]
for r in range(HQ_FIRST, HQ_LAST+1):
    i = r - HQ_FIRST
    zebra = PALE if i % 2 else None
    setcell(hq, f"A{r}", f'=IF(B{r}="","",ROW()-{HQ_FIRST-1})', INT, f(), zebra, CENTER)
    for col, fmt, al in [("B", None, LEFTC), ("C", DATE, CENTER), ("D", DATE, CENTER), ("E", None, LEFTC), ("F", None, CENTER), ("G", NUM2, RIGHT), ("T", None, LEFTC)]:
        setcell(hq, f"{col}{r}", None, fmt, f(color=INPUT_BLUE), zebra, al)
    setcell(hq, f"H{r}", f'=IF(B{r}="","",IFERROR(INDEX(Dashboard!$G$9:$G$12,MATCH(F{r},Dashboard!$F$9:$F$12,0)),1))', FX, f(color=LINK_GREEN), zebra, RIGHT)
    setcell(hq, f"I{r}", f'=IF(B{r}="","",G{r}*H{r})', TRY, f(), zebra, RIGHT)
    for col in "JKLMN":
        setcell(hq, f"{col}{r}", None, TRY, f(color=INPUT_BLUE), zebra, RIGHT)
    setcell(hq, f"O{r}", f'=IF(B{r}="","",SUM(J{r}:N{r}))', TRY, f(), zebra, RIGHT)
    setcell(hq, f"P{r}", f'=IF(B{r}="","",I{r}-O{r})', TRY, f(True), zebra, RIGHT)
    setcell(hq, f"Q{r}", f'=IF(OR(B{r}="",D{r}="",P{r}<=0),"",MAX(0,{REPORT_DATE}-D{r}))', INT, f(), zebra, CENTER)
    setcell(hq, f"R{r}", f'=IF(OR(B{r}="",P{r}<=0),"",IF(Q{r}=0,"Not due",IF(Q{r}<=30,"1-30 days",IF(Q{r}<=60,"31-60 days",IF(Q{r}<=90,"61-90 days","90+ days")))))', None, f(), zebra, CENTER)
    setcell(hq, f"S{r}", f'=IF(B{r}="","",IF(P{r}<=0,"Settled",IF(AND(Q{r}<>"",Q{r}>0),"Overdue","Open")))', None, f(True), zebra, CENTER)
    if i < len(hq_sample):
        s = hq_sample[i]
        hq[f"B{r}"] = s[0]; hq[f"C{r}"] = s[1]; hq[f"D{r}"] = s[2]; hq[f"E{r}"] = s[3]; hq[f"F{r}"] = s[4]; hq[f"G{r}"] = s[5]
        hq[f"T{r}"] = "Sample row – replace with live data"
TH = HQ_LAST + 1
setcell(hq, f"B{TH}", "TOTAL", None, f(True, NAVY, 11), LIGHT, LEFTC)
for col in "ACDEFGHRT":
    setcell(hq, f"{col}{TH}", None, None, f(), LIGHT)
for col in "IJKLMNOP":
    setcell(hq, f"{col}{TH}", f"=SUM({col}{HQ_FIRST}:{col}{HQ_LAST})", TRY, f(True, NAVY, 11), LIGHT, RIGHT)
setcell(hq, f"Q{TH}", f'=COUNTIF(S{HQ_FIRST}:S{HQ_LAST},"Overdue")', INT, f(True, "C00000", 11), LIGHT, CENTER)
setcell(hq, f"S{TH}", "overdue", None, f(True, "C00000", 9), LIGHT, CENTER)
for col in range(1, 21):
    hq.cell(row=TH, column=col).border = TOP2
# currency subtotals
setcell(hq, f"E{TH+2}", "Outstanding by currency (FCY)", None, f(True, NAVY), border=False)
for k, cur in enumerate(["GBP", "EUR", "USD", "TRY"]):
    rr = TH+3+k
    setcell(hq, f"E{rr}", cur, None, f(), None, LEFTC)
    setcell(hq, f"F{rr}", f'=SUMIF(F{HQ_FIRST}:F{HQ_LAST},E{rr},G{HQ_FIRST}:G{HQ_LAST})-SUMPRODUCT((F{HQ_FIRST}:F{HQ_LAST}=E{rr})*(O{HQ_FIRST}:O{HQ_LAST}<>"")*IFERROR(O{HQ_FIRST}:O{HQ_LAST}/H{HQ_FIRST}:H{HQ_LAST},0))', NUM2, f(), None, RIGHT)
    hq.merge_cells(f"F{rr}:G{rr}")
hq[f"F{TH+3}"].comment = Comment("Invoiced FCY less settlements converted back at the invoice FX rate.", "Finance")
hq.freeze_panes = f"C{HQ_FIRST}"
hq.auto_filter.ref = f"A{HR}:T{HQ_LAST}"
dvcur = DataValidation(type="list", formula1='"GBP,EUR,USD,TRY"', allow_blank=True, error="Choose GBP, EUR, USD or TRY", errorTitle="Currency")
hq.add_data_validation(dvcur); dvcur.add(f"F{HQ_FIRST}:F{HQ_LAST}")
dvd3 = DataValidation(type="date", operator="greaterThan", formula1="36526", allow_blank=True); hq.add_data_validation(dvd3); dvd3.add(f"C{HQ_FIRST}:D{HQ_LAST}")
rng = f"S{HQ_FIRST}:S{HQ_LAST}"
hq.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Overdue"'], fill=fill(RED), font=Font(name=FONT, bold=True, color="9C0006")))
hq.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Settled"'], fill=fill(GREEN), font=Font(name=FONT, bold=True, color="006100")))
hq.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Open"'], fill=fill(AMBER), font=Font(name=FONT, bold=True, color="9C5700")))
hq.conditional_formatting.add(f"R{HQ_FIRST}:R{HQ_LAST}", CellIsRule(operator="equal", formula=['"90+ days"'], fill=fill(RED), font=Font(name=FONT, bold=True, color="9C0006")))

# =====================================================================
# BANK BALANCES
# =====================================================================
title(bk, "BANK & CASH BALANCES", "Opening liquidity as at the report date. Foreign currency balances convert to ₺ using the Dashboard FX rates.", 10)
bk_cols = ["No", "Bank", "Account Name / IBAN (last 4)", "Account Type", "Currency", "Balance (FCY)", "FX Rate (to ₺)", "Balance (₺)", "Available for Operations?", "Notes"]
bk_w = [5, 24, 30, 16, 10, 18, 12, 18, 14, 30]
for i, (h, w) in enumerate(zip(bk_cols, bk_w), 1):
    hdr(bk, 4, i, h, w)
bk.row_dimensions[4].height = 32
BK_FIRST, BK_LAST = 5, 24
bk_sample = [
    ("Garanti BBVA", "Operating account – ****4821", "Current", "TRY", 1250000, "Yes"),
    ("İş Bankası", "FX account – ****1177", "Current", "EUR", 15000, "Yes"),
    ("Garanti BBVA", "Overnight deposit – ****4839", "Deposit", "TRY", 500000, "No"),
    ("Petty cash", "Office safe", "Cash", "TRY", 8500, "Yes"),
]
for r in range(BK_FIRST, BK_LAST+1):
    i = r - BK_FIRST
    zebra = PALE if i % 2 else None
    setcell(bk, f"A{r}", f'=IF(B{r}="","",ROW()-{BK_FIRST-1})', INT, f(), zebra, CENTER)
    for col, fmt, al in [("B", None, LEFTC), ("C", None, LEFTC), ("D", None, CENTER), ("E", None, CENTER), ("F", NUM2, RIGHT), ("I", None, CENTER), ("J", None, LEFTC)]:
        setcell(bk, f"{col}{r}", None, fmt, f(color=INPUT_BLUE), zebra, al)
    setcell(bk, f"G{r}", f'=IF(B{r}="","",IFERROR(INDEX(Dashboard!$G$9:$G$12,MATCH(E{r},Dashboard!$F$9:$F$12,0)),1))', FX, f(color=LINK_GREEN), zebra, RIGHT)
    setcell(bk, f"H{r}", f'=IF(B{r}="","",F{r}*G{r})', TRY, f(True), zebra, RIGHT)
    if i < len(bk_sample):
        s = bk_sample[i]
        bk[f"B{r}"], bk[f"C{r}"], bk[f"D{r}"], bk[f"E{r}"], bk[f"F{r}"], bk[f"I{r}"] = s
        bk[f"J{r}"] = "Sample row – replace with live data"
TB = BK_LAST + 1
setcell(bk, f"B{TB}", "TOTAL BANK & CASH", None, f(True, NAVY, 11), LIGHT, LEFTC)
for col in "ACDEFGIJ":
    setcell(bk, f"{col}{TB}", None, None, f(), LIGHT)
setcell(bk, f"H{TB}", f"=SUM(H{BK_FIRST}:H{BK_LAST})", TRY, f(True, NAVY, 11), LIGHT, RIGHT)
for col in range(1, 11):
    bk.cell(row=TB, column=col).border = TOP2
setcell(bk, f"B{TB+1}", "of which available for operations", None, f(italic=True), None, LEFTC)
setcell(bk, f"H{TB+1}", f'=SUMIF(I{BK_FIRST}:I{BK_LAST},"Yes",H{BK_FIRST}:H{BK_LAST})', TRY, f(True), None, RIGHT)
setcell(bk, f"B{TB+2}", "of which restricted / deposits", None, f(italic=True), None, LEFTC)
setcell(bk, f"H{TB+2}", f"=H{TB}-H{TB+1}", TRY, f(True), None, RIGHT)
setcell(bk, f"B{TB+4}", "Balances by currency (FCY)", None, f(True, NAVY), border=False)
for k, cur in enumerate(["TRY", "GBP", "EUR", "USD"]):
    rr = TB+5+k
    setcell(bk, f"B{rr}", cur, None, f(), None, LEFTC)
    setcell(bk, f"C{rr}", f'=SUMIF(E{BK_FIRST}:E{BK_LAST},B{rr},F{BK_FIRST}:F{BK_LAST})', NUM2, f(), None, RIGHT)
    setcell(bk, f"D{rr}", f'=SUMIF(E{BK_FIRST}:E{BK_LAST},B{rr},H{BK_FIRST}:H{BK_LAST})', TRY, f(), None, RIGHT)
bk.cell(row=TB+4, column=3, value="Amount (FCY)").font = f(True, NAVY, 9)
bk.cell(row=TB+4, column=4, value="Equivalent (₺)").font = f(True, NAVY, 9)
bk.freeze_panes = f"C{BK_FIRST}"
dvt = DataValidation(type="list", formula1='"Current,Deposit,Credit Line,Cash,Other"', allow_blank=True); bk.add_data_validation(dvt); dvt.add(f"D{BK_FIRST}:D{BK_LAST}")
dvc2 = DataValidation(type="list", formula1='"TRY,GBP,EUR,USD"', allow_blank=True); bk.add_data_validation(dvc2); dvc2.add(f"E{BK_FIRST}:E{BK_LAST}")
dvy = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True); bk.add_data_validation(dvy); dvy.add(f"I{BK_FIRST}:I{BK_LAST}")

# =====================================================================
# DASHBOARD
# =====================================================================
d = dash
title(d, "CASH FLOW STATEMENT – EXECUTIVE DASHBOARD", "All figures in Turkish Lira (₺) unless stated. Every number on this page is a formula driven by the four register tabs.", 12)
for col, w in zip("ABCDEFGHIJKL", [3, 30, 18, 3, 3, 26, 18, 18, 18, 18, 18, 3]):
    d.column_dimensions[col].width = w

# --- Inputs block
d.merge_cells("B4:C4"); setcell(d, "B4", "REPORT SETTINGS  (blue = input)", None, f(True, "FFFFFF"), MID, LEFTC)
d["C4"].fill = fill(MID)

rows = [("Company / Entity", "Company Name Ltd. – Türkiye", None),
        ("Report Period", "August 2026", None),
        ("Prepared By", "Finance Department", None),
        ("Report Date (as at)", dt.date(2026, 8, 31), DATE)]
for k, (lab, val, fmt) in enumerate(rows):
    r = 5 + k
    setcell(d, f"B{r}", lab, None, f(True), GREY, LEFTC)
    setcell(d, f"C{r}", val, fmt, f(color=INPUT_BLUE), YELLOW if fmt else None, CENTER if fmt else LEFTC)
d["C8"].comment = Comment("Report date drives every 'Days Overdue' and status calculation. User input.", "Finance")
setcell(d, "B9", "WEEK START DATES", None, f(True, "FFFFFF"), MID, LEFTC); setcell(d, "C9", "Start date", None, f(True, "FFFFFF"), MID, CENTER)
for k in range(5):
    r = 10 + k
    setcell(d, f"B{r}", f"Week {k+1}", None, f(True), GREY, LEFTC)
    if k == 0:
        setcell(d, f"C{r}", dt.date(2026, 8, 3), DATE, f(color=INPUT_BLUE), YELLOW, CENTER)
    else:
        setcell(d, f"C{r}", f"=C{r-1}+7", DATE, f(), None, CENTER)
d["C10"].comment = Comment("Enter the Monday of week 1. Weeks 2-5 are calculated (+7 days).", "Finance")

# FX table
d.merge_cells("F7:G7"); setcell(d, "F7", "FX RATES TO ₺ (blue = input)", None, f(True, "FFFFFF"), MID, LEFTC); d["G7"].fill = fill(MID)
setcell(d, "F8", "Currency", None, f(True), GREY, CENTER); setcell(d, "G8", "1 unit = ₺", None, f(True), GREY, CENTER)
for k, (cur, rate) in enumerate([("GBP", 55.20), ("EUR", 47.80), ("USD", 41.30), ("TRY", 1.0)]):
    r = 9 + k
    setcell(d, f"F{r}", cur, None, f(True), None, CENTER)
    setcell(d, f"G{r}", rate, FX, f(color=INPUT_BLUE), YELLOW if cur != "TRY" else None, RIGHT)
d["G9"].comment = Comment("Indicative rates entered by the user – update to the CBRT / bank rate at the report date. TRY is fixed at 1.", "Finance")
d["G12"].font = f()

# Legend
setcell(d, "I7", "LEGEND", None, f(True, "FFFFFF"), MID, LEFTC); d.merge_cells("I7:K7"); d["J7"].fill = fill(MID); d["K7"].fill = fill(MID)
leg = [("Blue text / yellow fill", "User input", INPUT_BLUE, YELLOW),
       ("Black text", "Formula – do not edit", "000000", None),
       ("Green text", "Link to another tab", LINK_GREEN, None),
       ("Red status", "Overdue – action required", "9C0006", RED),
       ("Green status", "Collected / Paid / Settled", "006100", GREEN)]
for k, (a, b, colr, fl) in enumerate(leg):
    r = 8 + k
    setcell(d, f"I{r}", a, None, f(True, colr), fl, LEFTC); d.merge_cells(f"J{r}:K{r}")
    setcell(d, f"J{r}", b, None, f(), None, LEFTC); d[f"K{r}"].border = BORDER
setcell(d, "I13", "Sample rows in each tab are examples – overwrite them with live data.", None, f(italic=True, size=9), border=False)

# --- KPI tiles row 16-18
d.merge_cells("B16:K16"); setcell(d, "B16", "KEY LIQUIDITY INDICATORS (₺)", None, f(True, "FFFFFF", 11), NAVY, LEFTC)
tiles = [
 ("B", "Bank & Cash Balance", f"='Bank Balances'!H{TB}", NAVY),
 ("C", "Receivables Outstanding", f"=Receivables!M{TR}", "375623"),
 ("F", "Payables Outstanding", f"=Payables!N{TP}", "C00000"),
 ("G", "UK HQ Outstanding", f"='UK HQ Invoices'!P{TH}", "C00000"),
 ("H", "Total Liabilities", "=F18+G18", "7F6000"),
 ("I", "Net Cash Position", "=B18+C18-H18", NAVY),
 ("J", "Liquidity Ratio", '=IF(H18=0,"n/a",(B18+C18)/H18)', "7F6000"),
]
for col, lab, frm, colr in tiles:
    setcell(d, f"{col}17", lab, None, f(True, "FFFFFF", 9), MID, CENTER)
    setcell(d, f"{col}18", frm, '0.00"x"' if "Ratio" in lab else TRY, Font(name=FONT, bold=True, size=12, color=colr), LIGHT, CENTER)
    d.row_dimensions[18].height = 28
d["I18"].comment = Comment("Net Cash Position = Bank & Cash + Receivables outstanding – Payables outstanding – UK HQ outstanding.", "Finance")
d["J18"].comment = Comment("Liquidity Ratio = (Bank & Cash + Receivables) / (Payables + UK HQ). Above 1.0x means liquid assets cover all obligations.", "Finance")
d.conditional_formatting.add("I18", CellIsRule(operator="lessThan", formula=["0"], font=Font(name=FONT, bold=True, size=12, color="C00000"), fill=fill(RED)))
d.conditional_formatting.add("J18", CellIsRule(operator="lessThan", formula=["1"], font=Font(name=FONT, bold=True, size=12, color="C00000"), fill=fill(RED)))
d.merge_cells("D17:E18")

# --- Weekly cash flow projection rows 20-28
d.merge_cells("B20:K20"); setcell(d, "B20", "5-WEEK CASH FLOW PROJECTION (₺)", None, f(True, "FFFFFF", 11), NAVY, LEFTC)
heads = ["Line item", "Week 1", "Week 2", "Week 3", "Week 4", "Week 5", "Total"]
# Layout: B label, F..J weeks, K total (C,D,E merged to B)
setcell(d, "B21", "Line item", None, f(True), GREY, LEFTC); d.merge_cells("B21:E21")
for k in range(5):
    c = d.cell(row=21, column=6+k, value=f"=C{10+k}"); c.number_format = DATE; c.font = f(True, LINK_GREEN, 9); c.fill = fill(GREY); c.alignment = CENTER; c.border = BORDER
setcell(d, "K21", "Total", None, f(True), GREY, CENTER)
d.merge_cells("B22:E22"); setcell(d, "B22", "Week", None, f(True), GREY, LEFTC)
for k in range(5):
    setcell(d, f"{L(6+k)}22", f"Week {k+1}", None, f(True), GREY, CENTER)
setcell(d, "K22", "", None, f(), GREY)
lines = [
 (23, "Opening balance", "open"),
 (24, "(+) Customer collections", "ar"),
 (25, "(–) Supplier & operating payments", "ap"),
 (26, "(–) UK HQ settlements", "hq"),
 (27, "Net cash flow", "net"),
 (28, "Closing balance", "close"),
]
ar_wk = "GHIJK"; ap_wk = "HIJKL"; hq_wk = "JKLMN"
for r, lab, kind in lines:
    bold = kind in ("open", "net", "close")
    d.merge_cells(f"B{r}:E{r}")
    setcell(d, f"B{r}", lab, None, f(bold, NAVY if bold else "000000"), LIGHT if bold else None, LEFTC)
    for k in range(5):
        col = L(6+k)
        if kind == "open":
            frm = f"='Bank Balances'!H{TB}" if k == 0 else f"={L(5+k)}28"
        elif kind == "ar":  frm = f"=Receivables!{ar_wk[k]}{TR}"
        elif kind == "ap":  frm = f"=-Payables!{ap_wk[k]}{TP}"
        elif kind == "hq":  frm = f"=-'UK HQ Invoices'!{hq_wk[k]}{TH}"
        elif kind == "net": frm = f"=SUM({col}24:{col}26)"
        else:               frm = f"={col}23+{col}27"
        setcell(d, f"{col}{r}", frm, TRY, f(bold, LINK_GREEN if kind in ("ar","ap","hq") else "000000"), LIGHT if bold else None, RIGHT)
    if kind in ("ar", "ap", "hq", "net"):
        setcell(d, f"K{r}", f"=SUM(F{r}:J{r})", TRY, f(True), LIGHT if bold else None, RIGHT)
    elif kind == "open":
        setcell(d, f"K{r}", "=F23", TRY, f(True), LIGHT, RIGHT)
    else:
        setcell(d, f"K{r}", "=J28", TRY, f(True), LIGHT, RIGHT)
for col in range(2, 12):
    d.cell(row=28, column=col).border = TOP2
d.conditional_formatting.add("F28:K28", CellIsRule(operator="lessThan", formula=["0"], font=Font(name=FONT, bold=True, color="C00000"), fill=fill(RED)))
d.conditional_formatting.add("F27:K27", CellIsRule(operator="lessThan", formula=["0"], font=Font(name=FONT, bold=True, color="C00000")))
setcell(d, "B29", "Control: Closing week 5 = Bank & Cash + total net cash flow (must be 0)", None, f(italic=True, size=9), border=False)
setcell(d, "K29", "=J28-(B18+K27)", TRY, f(True, size=9), None, RIGHT)
setcell(d, "B30", "Note: the projection assumes only the collections / payments scheduled in the week columns of each tab occur. Unscheduled balances remain outstanding.", None, f(italic=True, size=9), border=False)

# --- Breakdowns row 32+
d.merge_cells("B32:K32"); setcell(d, "B32", "POSITION ANALYSIS", None, f(True, "FFFFFF", 11), NAVY, LEFTC)
# Receivables by status  (B33:C38)
setcell(d, "B33", "Receivables by status", None, f(True, "FFFFFF"), MID, LEFTC); setcell(d, "C33", "Outstanding (₺)", None, f(True, "FFFFFF"), MID, CENTER)
for k, st in enumerate(["Open", "Overdue", "Collected"]):
    r = 34 + k
    setcell(d, f"B{r}", st, None, f(), None, LEFTC)
    if st == "Collected":
        setcell(d, f"C{r}", f'=SUMIF(Receivables!P{AR_FIRST}:P{AR_LAST},"Collected",Receivables!L{AR_FIRST}:L{AR_LAST})', TRY, f(color=LINK_GREEN), None, RIGHT)
        d[f"C{r}"].comment = Comment("Collected shows the amount already received on fully-collected invoices.", "Finance")
    else:
        setcell(d, f"C{r}", f'=SUMIF(Receivables!P{AR_FIRST}:P{AR_LAST},"{st}",Receivables!M{AR_FIRST}:M{AR_LAST})', TRY, f(color=LINK_GREEN), None, RIGHT)
setcell(d, "B37", "Invoiced total", None, f(True), LIGHT, LEFTC); setcell(d, "C37", f"=Receivables!F{TR}", TRY, f(True, LINK_GREEN), LIGHT, RIGHT)
setcell(d, "B38", "Collection rate", None, f(True), LIGHT, LEFTC); setcell(d, "C38", f"=Receivables!N{TR}", PCT, f(True, LINK_GREEN), LIGHT, RIGHT)
setcell(d, "B39", "Overdue invoices (count)", None, f(True), LIGHT, LEFTC); setcell(d, "C39", f"=Receivables!O{TR}", INT, f(True, "C00000"), LIGHT, RIGHT)

# Payables by category (F33:G43)
setcell(d, "F33", "Payables by category", None, f(True, "FFFFFF"), MID, LEFTC); setcell(d, "G33", "Outstanding (₺)", None, f(True, "FFFFFF"), MID, CENTER)
for k, cat in enumerate(CATS):
    r = 34 + k
    setcell(d, f"F{r}", cat, None, f(), None, LEFTC)
    setcell(d, f"G{r}", f'=SUMIF(Payables!C{AP_FIRST}:C{AP_LAST},F{r},Payables!N{AP_FIRST}:N{AP_LAST})', TRY, f(color=LINK_GREEN), None, RIGHT)
r = 34 + len(CATS)
setcell(d, f"F{r}", "Total outstanding", None, f(True), LIGHT, LEFTC); setcell(d, f"G{r}", f"=SUM(G34:G{r-1})", TRY, f(True), LIGHT, RIGHT)
setcell(d, f"F{r+1}", "Check vs Payables tab (must be 0)", None, f(italic=True, size=9), None, LEFTC); setcell(d, f"G{r+1}", f"=G{r}-Payables!N{TP}", TRY, f(size=9), None, RIGHT)
setcell(d, f"F{r+2}", "Overdue payables (count)", None, f(True), LIGHT, LEFTC); setcell(d, f"G{r+2}", f"=Payables!P{TP}", INT, f(True, "C00000"), LIGHT, RIGHT)

# UK HQ aging (I33:K39)
setcell(d, "I33", "UK HQ aging", None, f(True, "FFFFFF"), MID, LEFTC); setcell(d, "J33", "Outstanding (₺)", None, f(True, "FFFFFF"), MID, CENTER); setcell(d, "K33", "% of total", None, f(True, "FFFFFF"), MID, CENTER)
buckets = ["Not due", "1-30 days", "31-60 days", "61-90 days", "90+ days"]
for k, b in enumerate(buckets):
    r = 34 + k
    setcell(d, f"I{r}", b, None, f(), None, LEFTC)
    setcell(d, f"J{r}", f"=SUMIF('UK HQ Invoices'!R{HQ_FIRST}:R{HQ_LAST},I{r},'UK HQ Invoices'!P{HQ_FIRST}:P{HQ_LAST})", TRY, f(color=LINK_GREEN), None, RIGHT)
    setcell(d, f"K{r}", f'=IF($J$39=0,"",J{r}/$J$39)', PCT, f(), None, RIGHT)
setcell(d, "I39", "Total outstanding", None, f(True), LIGHT, LEFTC); setcell(d, "J39", "=SUM(J34:J38)", TRY, f(True), LIGHT, RIGHT); setcell(d, "K39", '=IF(J39=0,"",1)', PCT, f(True), LIGHT, RIGHT)
setcell(d, "I40", "Overdue HQ invoices (count)", None, f(True), LIGHT, LEFTC); setcell(d, "J40", f"='UK HQ Invoices'!Q{TH}", INT, f(True, "C00000"), LIGHT, RIGHT); setcell(d, "K40", "", None, f(), LIGHT)
d.conditional_formatting.add("J38", CellIsRule(operator="greaterThan", formula=["0"], font=Font(name=FONT, bold=True, color="C00000"), fill=fill(RED)))

# Bank by currency (B41:C46)
setcell(d, "B41", "Bank & cash by currency", None, f(True, "FFFFFF"), MID, LEFTC); setcell(d, "C41", "Equivalent (₺)", None, f(True, "FFFFFF"), MID, CENTER)
for k, cur in enumerate(["TRY", "GBP", "EUR", "USD"]):
    r = 42 + k
    setcell(d, f"B{r}", cur, None, f(), None, LEFTC)
    setcell(d, f"C{r}", f"=SUMIF('Bank Balances'!E{BK_FIRST}:E{BK_LAST},B{r},'Bank Balances'!H{BK_FIRST}:H{BK_LAST})", TRY, f(color=LINK_GREEN), None, RIGHT)
setcell(d, "B46", "Available for operations", None, f(True), LIGHT, LEFTC); setcell(d, "C46", f"='Bank Balances'!H{TB+1}", TRY, f(True, LINK_GREEN), LIGHT, RIGHT)

# --- Charts
ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"; ch.title = "Weekly cash inflows vs outflows (₺)"
ch.y_axis.title = "₺"; ch.x_axis.title = None; ch.height = 8; ch.width = 18
data = Reference(d, min_col=2, min_row=24, max_col=10, max_row=26)   # B..J rows 24-26 (labels in B via merged)
ch.add_data(data, from_rows=True, titles_from_data=True)
ch.set_categories(Reference(d, min_col=6, min_row=22, max_col=10, max_row=22))
# Remove C..E blank columns from the series: rebuild with explicit refs
ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"; ch.title = "Weekly cash inflows vs outflows (₺)"; ch.height = 8; ch.width = 18
for r, name in [(24, "Collections"), (25, "Payments"), (26, "UK HQ")]:
    from openpyxl.chart import Series
    s = Series(Reference(d, min_col=6, min_row=r, max_col=10, max_row=r), title=name)
    ch.series.append(s)
ch.set_categories(Reference(d, min_col=6, min_row=22, max_col=10, max_row=22))
ch.y_axis.numFmt = '#,##0'
d.add_chart(ch, "B49")

lc = LineChart(); lc.title = "Projected closing bank balance (₺)"; lc.height = 8; lc.width = 14
from openpyxl.chart import Series
s = Series(Reference(d, min_col=6, min_row=28, max_col=10, max_row=28), title="Closing balance"); lc.series.append(s)
lc.set_categories(Reference(d, min_col=6, min_row=22, max_col=10, max_row=22)); lc.y_axis.numFmt = '#,##0'
d.add_chart(lc, "H49")

d.freeze_panes = "A4"
for ws in wb.worksheets:
    ws.sheet_properties.tabColor = {"Dashboard": NAVY, "Receivables": "375623", "Payables": "C00000", "UK HQ Invoices": "7F6000", "Bank Balances": "2F5496"}[ws.title]
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:5" if ws.title != "Dashboard" else "1:3"
    ws.oddFooter.center.text = "&A – Page &P of &N"; ws.oddHeader.right.text = "Confidential"

wb.save(OUT); print("saved", OUT)
